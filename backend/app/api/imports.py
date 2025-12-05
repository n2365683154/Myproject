"""
楚然智考系统 - 题库导入API路由
"""
import os
import uuid
import aiofiles
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Query
from sqlalchemy.orm import Session
from typing import Optional

from app.database import get_db
from app.config import settings
from app.services.import_service import ImportService
from app.schemas.question import ImportResult, QuestionBankResponse, QuestionBankListResponse
from app.api.deps import get_current_user, requires_permission
from app.models.user import User
from app.models.permission import PermissionCode


router = APIRouter()


async def save_upload_file(upload_file: UploadFile, allowed_extensions: list) -> str:
    """
    保存上传文件
    返回文件路径
    """
    # 检查文件扩展名
    ext = os.path.splitext(upload_file.filename)[1].lower()
    if ext not in allowed_extensions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"不支持的文件格式，仅支持: {', '.join(allowed_extensions)}"
        )
    
    # 检查文件大小
    content = await upload_file.read()
    if len(content) > settings.MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"文件大小超过限制({settings.MAX_UPLOAD_SIZE // 1024 // 1024}MB)"
        )
    
    # 创建上传目录
    upload_dir = settings.UPLOAD_DIR
    os.makedirs(upload_dir, exist_ok=True)
    
    # 生成唯一文件名
    filename = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(upload_dir, filename)
    
    # 保存文件
    async with aiofiles.open(file_path, 'wb') as f:
        await f.write(content)
    
    return file_path


@router.post("/excel", response_model=ImportResult, summary="Excel导入题库")
async def import_from_excel(
    file: UploadFile = File(..., description="Excel文件(.xlsx)"),
    db: Session = Depends(get_db),
    current_user: User = requires_permission(PermissionCode.QUESTION_IMPORT)
):
    """
    从Excel文件导入题库
    
    Excel格式要求：
    - 第一行为表头
    - 列顺序：题型 | 题干 | 选项 | 答案 | 解析 | 知识点 | 难度
    - 题型：单选题/多选题/判断题/填空题/简答题
    - 选项格式：A.选项A B.选项B C.选项C D.选项D
    - 难度：简单/中等/困难
    """
    file_path = await save_upload_file(file, ['.xlsx', '.xls'])
    
    try:
        import_service = ImportService(db)
        result = import_service.import_from_excel(file_path, current_user.id)
        return result
    finally:
        # 清理临时文件
        if os.path.exists(file_path):
            os.remove(file_path)


@router.post("/word", response_model=ImportResult, summary="Word导入题库")
async def import_from_word(
    file: UploadFile = File(..., description="Word文件(.docx)"),
    bank_name: str = Form(..., description="题库名称"),
    db: Session = Depends(get_db),
    current_user: User = requires_permission(PermissionCode.QUESTION_IMPORT)
):
    """
    从Word文件导入题库
    
    Word格式要求：
    - 使用【选择题】、【多选题】、【判断题】、【填空题】、【简答题】标题区分题型
    - 题目格式：1. 题干内容
    - 选项格式：A. 选项内容
    - 答案格式：答案：A
    - 解析格式：解析：解析内容
    """
    if not bank_name or not bank_name.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="题库名称不能为空"
        )
    
    file_path = await save_upload_file(file, ['.docx', '.doc'])
    
    try:
        import_service = ImportService(db)
        result = import_service.import_from_word(file_path, bank_name.strip(), current_user.id)
        return result
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@router.post("/ocr", response_model=ImportResult, summary="图片OCR导入题库")
async def import_from_image(
    file: UploadFile = File(..., description="图片文件(.jpg/.png)"),
    db: Session = Depends(get_db),
    current_user: User = requires_permission(PermissionCode.QUESTION_IMPORT)
):
    """
    从图片OCR识别导入题库
    
    支持格式：.jpg, .jpeg, .png
    
    注意事项：
    - 图片清晰度会影响识别准确率
    - 建议使用打印体文字
    - 识别后建议人工校验
    """
    file_path = await save_upload_file(file, ['.jpg', '.jpeg', '.png'])
    
    try:
        import_service = ImportService(db)
        result = import_service.import_from_image(file_path, current_user.id)
        return result
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@router.post("/ocr/preview", summary="OCR预览（不入库）")
async def ocr_preview(
    file: UploadFile = File(..., description="图片文件(.jpg/.png)"),
    db: Session = Depends(get_db),
    current_user: User = requires_permission(PermissionCode.QUESTION_IMPORT)
):
    """
    OCR识别预览
    
    返回识别结果供用户确认和修改，不直接入库
    """
    file_path = await save_upload_file(file, ['.jpg', '.jpeg', '.png'])
    
    try:
        import_service = ImportService(db)
        result = import_service.ocr_preview(file_path)
        return result
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@router.post("/pdf", response_model=ImportResult, summary="PDF导入题库")
async def import_from_pdf(
    file: UploadFile = File(..., description="PDF文件(.pdf)"),
    bank_name: str = Form(..., description="题库名称"),
    db: Session = Depends(get_db),
    current_user: User = requires_permission(PermissionCode.QUESTION_IMPORT)
):
    """
    从PDF文件导入题库
    
    PDF格式要求：
    - 支持标准文本PDF（非扫描件效果更好）
    - 使用【选择题】、【多选题】、【判断题】等标题区分题型
    - 题目格式：1. 题干内容 或 1、题干内容
    - 选项格式：A. 选项内容 或 A、选项内容
    - 答案格式：题干中包含（A）或 答案：A
    """
    if not bank_name or not bank_name.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="题库名称不能为空"
        )
    
    file_path = await save_upload_file(file, ['.pdf'])
    
    try:
        import_service = ImportService(db)
        result = import_service.import_from_pdf(file_path, bank_name.strip(), current_user.id)
        return result
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@router.post("/pdf/preview", summary="PDF预览（不入库）")
async def pdf_preview(
    file: UploadFile = File(..., description="PDF文件(.pdf)"),
    db: Session = Depends(get_db),
    current_user: User = requires_permission(PermissionCode.QUESTION_IMPORT)
):
    """
    PDF识别预览
    
    返回识别结果供用户确认和修改，不直接入库
    """
    file_path = await save_upload_file(file, ['.pdf'])
    
    try:
        import_service = ImportService(db)
        result = import_service.pdf_preview(file_path)
        return result
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@router.get("/template/excel", summary="下载Excel模板")
async def download_excel_template():
    """
    下载Excel导入模板
    """
    from fastapi.responses import FileResponse
    
    template_path = os.path.join(settings.UPLOAD_DIR, "templates", "question_template.xlsx")
    
    # 如果模板不存在，创建一个
    if not os.path.exists(template_path):
        os.makedirs(os.path.dirname(template_path), exist_ok=True)
        
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "题库导入模板"
        
        # 表头
        headers = ["题型", "题干", "选项", "答案", "解析", "知识点", "难度"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 示例数据
        examples = [
            ["单选题", "以下哪个是Python的关键字？", "A.class B.function C.method D.define", "A", "class是Python的关键字，用于定义类", "Python基础", "简单"],
            ["多选题", "以下哪些是Python的数据类型？", "A.int B.float C.string D.list", "ABCD", "Python支持多种数据类型", "Python基础", "中等"],
            ["判断题", "Python是一种解释型语言", "", "对", "Python代码在运行时由解释器逐行解释执行", "Python基础", "简单"],
            ["填空题", "Python中用____关键字定义函数", "", "def", "def是定义函数的关键字", "Python基础", "简单"],
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        
        wb.save(template_path)
    
    return FileResponse(
        template_path,
        filename="题库导入模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==================== 题库管理 ====================

@router.get("/banks", response_model=QuestionBankListResponse, summary="获取题库列表")
async def get_banks(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取所有题库列表"""
    import_service = ImportService(db)
    result = import_service.get_banks(skip, limit)
    return result


@router.get("/banks/{bank_id}", response_model=QuestionBankResponse, summary="获取题库详情")
async def get_bank(
    bank_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取单个题库详情"""
    import_service = ImportService(db)
    bank = import_service.get_bank(bank_id)
    if not bank:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="题库不存在"
        )
    return bank


@router.delete("/banks/{bank_id}", summary="删除题库")
async def delete_bank(
    bank_id: int,
    db: Session = Depends(get_db),
    current_user: User = requires_permission(PermissionCode.QUESTION_DELETE)
):
    """删除题库及其所有题目"""
    import_service = ImportService(db)
    success = import_service.delete_bank(bank_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="题库不存在"
        )
    return {"message": "删除成功"}
